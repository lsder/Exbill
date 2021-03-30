import openpyxl,os,csv

#Excel中的Sheet
class sheet:
    def __init__(self,account):
        self.file_name=account.file
        self.sheet_name=account.name
        self.endrow=account.csv_endrow
        self.startrow=account.csv_startrow
        #判断文件是否存在
        if not os.path.isfile(self.file_name):
            wb1=openpyxl.Workbook()
            wb1.save(self.file_name)
        self.wb = openpyxl.load_workbook(self.file_name)
        #判断sheet是否存在
        if self.sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(self.sheet_name, 0)
        self.sheet = self.wb[self.sheet_name]   

    def __init__(self,file_name,sheet_name,start_row=0,end_row=0):
        self.file_name=file_name
        self.sheet_name=sheet_name
        self.endrow=end_row
        self.startrow=start_row
        #判断文件是否存在
        if not os.path.isfile(self.file_name):
            wb1=openpyxl.Workbook()
            wb1.save(self.file_name)
        self.wb = openpyxl.load_workbook(self.file_name)
        #判断sheet是否存在
        if self.sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(self.sheet_name, 0)
        self.sheet = self.wb[self.sheet_name]   
     
    def get_data_end(self):
        '''获取数据最后一行的位置'''
        return self.sheet.max_row-self.endrow

    def get_sheet_data(self):
        '''获取sheet中的数据行,返回字典'''
        rows=[]
        for i,row in enumerate(self.sheet.values):
            if i>self.startrow and i<self.sheet.max_row-self.endrow:
                sheet_data = ['','','','','','','','','','','','','','','','','']
                for i,j in enumerate(row):
                    #print(len(row))
                    #sheet_data[str(self.sheet[self.startrow][i].value).strip()]= str(row[i]).strip()
                    sheet_data[i]= str(j).strip()
                rows.append(sheet_data)
        return rows

    def get_id_list(self):
        '''获取sheet中的关键字  与插入位置'''
        rows=[]
        for i,row in enumerate(self.sheet.values):
            if i>self.startrow and i<self.sheet.max_row-self.endrow:
                rows.append(row[0])
        return rows
    
    def insert_row(self,pos,dat,quchong=True):
        '''插入一行'''
        if quchong and len(self.get_sheet_data()) != 0:
            for row in enumerate(self.get_sheet_data()):#遍历输入数据
                if row[1]== dat:return
        self.sheet.insert_rows(pos, amount=1)#插入新行
        for j,cell in enumerate(dat):
            self.sheet.cell(row=pos, column=j+1, value=cell)
   
    def insert_rows(self,pos,dat,quchong=True):
        '''插入多行'''
        for i,row_dat in enumerate(dat):
            
            if quchong and len(self.get_sheet_data()) != 0:
                for row in enumerate(self.get_sheet_data()):#遍历输入数据
                    if row[1]== row_dat:return
            self.sheet.insert_rows(pos+i, amount=1)#插入新行
            for j,cell in enumerate(row_dat):
                self.sheet.cell(row=pos+i, column=j+1, value=cell)

   
    def import_data_from_csv(self,csv_obj):
        '''导入数据到sheet ,数据为csv_loader对象'''
        if  self.sheet.max_row>(csv_obj.head_count+csv_obj.tail_count):#不是第一次导入
            start=self.get_data_end()+1
            id_list=self.get_id_list()#已经有的数据id表
            index=0
            for i,row in enumerate(csv_obj.data):#遍历输入数据
                if(row[0] not in id_list ):#去重
                        self.insert_row(start+index,row)
                        index=index+1

            print('插入',index,'条数据')
        else:#copy
            self.insert_rows(1,csv_obj.rows)
            print('插入',csv_obj.row_count,'条数据')
        self.wb.save(self.file_name)#保存数据

if __name__ =="__main__":
    import sys
    home='/'.join(sys.argv[0].split('/')[:-2])
    sys.path.append(home)
    from src.tool import csv_loader
    from src.account import account
    alipay=account(home+"/data/temp/alipay_record_20210103_1339_1.csv",'Alipay',5,7)
    fil=csv_loader.csv_file(alipay)
    alipay=sheet('tt.xlsx','alipay',start_row=5,end_row=7)
    alipay.import_data_from_csv(fil)